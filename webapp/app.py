"""
FastAPI 入口 —— 站群内容生产后台。

功能：
- 多站管理：新建 / 切换 / 编辑配置 / 删除
- 上传 Excel 导入关键词
- 一键运行生成（后台线程），实时看进度和日志
- 查看每条关键词的状态与发布链接

本地启动：  uvicorn webapp.app:app --reload
Render 启动：uvicorn webapp.app:app --host 0.0.0.0 --port $PORT
"""
import io
import os
from typing import Optional

# 注意：pandas 很占内存(~100MB+)，改为「用到 Excel 时才 import」，
# 避免 Web 启动/跑任务时白白占用，超出 Render 512MB 被 OOM。
from fastapi import FastAPI, Form, Request, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

import prompts as _prompts
from webapp import db
from webapp.pipeline import jobs

db.init_db()
# 服务启动自愈：把上次崩溃/重启残留的 processing 退回 pending
_orphaned = db.reset_orphaned_processing()
if _orphaned:
    print(f"[startup] 已把 {_orphaned} 条残留 processing 退回 pending")

BASE = os.path.dirname(__file__)
templates = Jinja2Templates(directory=os.path.join(BASE, "templates"))

app = FastAPI(title="站群内容生产后台")
app.mount("/static", StaticFiles(directory=os.path.join(BASE, "static")), name="static")


# ---------------- 首页 / 站点列表 ----------------
@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    sites = db.list_sites()
    return templates.TemplateResponse(request, "sites.html", {"sites": sites})


@app.get("/sites/new", response_class=HTMLResponse)
def new_site_form(request: Request):
    return templates.TemplateResponse(
        request, "site_edit.html",
        {"site": None, "fields": db.SITE_FIELDS, "defaults": db.SITE_FIELDS,
         "sites": db.list_sites()},
    )


@app.post("/sites")
async def create_site(request: Request):
    form = await request.form()
    name = (form.get("name") or "未命名站点").strip()
    # prompt 字段由专门的表单管理，这里不碰（新站用默认空 = 内置模板）
    values = {k: form.get(k, db.SITE_FIELDS[k]) for k in db.SITE_FIELDS if k not in db.PROMPT_FIELDS}
    values = _coerce(values, form)
    site_id = db.create_site(name, values)
    return RedirectResponse(f"/sites/{site_id}", status_code=303)


@app.get("/sites/{site_id}", response_class=HTMLResponse)
def site_detail(request: Request, site_id: int):
    site = db.get_site(site_id)
    if not site:
        return RedirectResponse("/", status_code=303)
    counts = db.keyword_counts(site_id)

    # 队列分页（避免一次渲染几千行）
    PAGE_SIZE = 50
    kw_status = request.query_params.get("kw_status") or "all"
    try:
        kw_page = max(1, int(request.query_params.get("kw_page") or 1))
    except ValueError:
        kw_page = 1
    kw_total = db.count_keywords(site_id, kw_status)
    kw_pages = max(1, (kw_total + PAGE_SIZE - 1) // PAGE_SIZE)
    kw_page = min(kw_page, kw_pages)
    keywords = db.list_keywords(site_id, status=kw_status,
                                limit=PAGE_SIZE, offset=(kw_page - 1) * PAGE_SIZE)
    return templates.TemplateResponse(
        request, "site_detail.html",
        {
            "site": site,
            "fields": db.SITE_FIELDS,
            "keywords": keywords,
            "counts": counts,
            "kw_status": kw_status,
            "kw_page": kw_page,
            "kw_pages": kw_pages,
            "kw_total": kw_total,
            "kw_page_size": PAGE_SIZE,
            "sites": db.list_sites(),
            "running": jobs.is_running(site_id),
            "prompt_defaults": _prompts.DEFAULTS,
            "prompt_tokens": _prompts.PROMPT_TOKENS,
            "prompt_labels": _prompts.PROMPT_LABELS,
            "presets": db.list_presets(),
            "runs": db.list_runs(site_id, 20),
        },
    )


@app.get("/sites/{site_id}/runs/{run_id}")
def get_run_log(site_id: int, run_id: int):
    r = db.get_run(run_id)
    return JSONResponse({"log": (r or {}).get("log", "") or "（无日志）"})


@app.post("/sites/{site_id}")
async def update_site(request: Request, site_id: int):
    form = await request.form()
    name = (form.get("name") or "未命名站点").strip()
    # 只更新非 prompt 的配置字段，prompt 交给 /prompts 路由
    values = {k: form.get(k, db.SITE_FIELDS[k]) for k in db.SITE_FIELDS if k not in db.PROMPT_FIELDS}
    values = _coerce(values, form)
    db.update_site(site_id, name, values)
    return RedirectResponse(f"/sites/{site_id}?saved=1", status_code=303)


@app.post("/sites/{site_id}/prompts")
async def update_prompts(request: Request, site_id: int):
    form = await request.form()
    values = {k: (form.get(k) or "") for k in db.PROMPT_FIELDS}
    db.update_site(site_id, values=values)
    return RedirectResponse(f"/sites/{site_id}?mode=prompt&saved=1", status_code=303)


# ---------------- Prompt 存档（预设）----------------
@app.post("/presets")
async def save_preset(request: Request):
    form = await request.form()
    db.add_preset(form.get("name") or "未命名", form.get("ptype") or "outline", form.get("content") or "")
    return JSONResponse({"ok": True})


@app.post("/presets/{preset_id}/delete")
def del_preset(preset_id: int):
    db.delete_preset(preset_id)
    return JSONResponse({"ok": True})


@app.post("/sites/{site_id}/delete")
def delete_site(site_id: int):
    db.delete_site(site_id)
    return RedirectResponse("/", status_code=303)


# ---------------- Excel 解析（上传 / 插入共用）----------------
# 用 openpyxl 直接读，不再 import pandas（pandas+numpy 常驻 ~50MB，
# 在 512MB 容器里是 OOM 的主要元凶之一；openpyxl 只有 ~7MB）。
def _parse_keyword_excel(content: bytes) -> list:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if not header:
        wb.close()
        return []
    cols = {str(h).lower().strip(): i for i, h in enumerate(header) if h is not None}

    def pick(row, *names, default=""):
        for n in names:
            if n in cols:
                v = row[cols[n]] if cols[n] < len(row) else None
                if v is not None and str(v).strip() != "":
                    return v
        return default

    rows = []
    for r in it:
        if r is None:
            continue
        main_kw = str(pick(r, "main_keyword", "main keyword", "keyword") or "").strip()
        if not main_kw:
            continue
        try:
            wc = int(float(pick(r, "wordcounts", "wordcount", "words", default=0) or 0))
        except (TypeError, ValueError):
            wc = 0
        rows.append({
            "main_keyword": main_kw,
            "secondary_keyword": str(pick(r, "secondary_keyword", "secondary keyword") or "").strip(),
            "topic": str(pick(r, "topic") or "").strip(),
            "wordcounts": wc,
            "specific": str(pick(r, "specific", "special") or "").strip(),
            "categories": str(pick(r, "categories", "category") or "").strip(),
            "tags": str(pick(r, "tags", "tag") or "").strip(),
        })
    wb.close()
    return rows


def _rows_to_xlsx(cols: list, dict_rows: list) -> io.BytesIO:
    """用 openpyxl 写 xlsx（替代 pandas.to_excel）。"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(cols)
    for r in dict_rows:
        ws.append([r.get(c, "") for c in cols])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------- Excel 模板下载 ----------------
@app.get("/template.xlsx")
def download_template():
    cols = ["main_keyword", "secondary_keyword", "topic", "wordcounts", "specific", "category", "tags"]
    buf = _rows_to_xlsx(cols, [
        {"main_keyword": "exness minimum deposit", "secondary_keyword": "exness deposit",
         "topic": "exness minimum deposit", "wordcounts": 2500,
         "specific": "面向新手", "category": "Giáo dục", "tags": "exness,deposit"},
        {"main_keyword": "lowest spread forex broker", "secondary_keyword": "low spread broker",
         "topic": "lowest spread forex broker", "wordcounts": 2500,
         "specific": "", "category": "", "tags": ""},
    ])
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="keywords_template.xlsx"'},
    )


# ---------------- 排定任务：Excel 上传 ----------------
@app.post("/sites/{site_id}/upload")
async def upload_excel(site_id: int, file: UploadFile = File(...)):
    content = await file.read()
    try:
        rows = _parse_keyword_excel(content)
    except Exception as e:
        return JSONResponse({"ok": False, "msg": f"读取 Excel 失败: {e}"}, status_code=400)
    if not rows:
        return JSONResponse({"ok": False, "msg": "没有解析到有效行（至少要有 main_keyword）"}, status_code=400)
    n = db.add_keywords(site_id, rows)
    return RedirectResponse(f"/sites/{site_id}?added={n}", status_code=303)


# ---------------- 插入任务：Excel 批量插入（可选立即生成）----------------
@app.post("/sites/{site_id}/insert_excel")
async def insert_excel(site_id: int, file: UploadFile = File(...), run_now: str = Form(None)):
    content = await file.read()
    try:
        rows = _parse_keyword_excel(content)
    except Exception as e:
        return JSONResponse({"ok": False, "msg": f"读取 Excel 失败: {e}"}, status_code=400)
    if not rows:
        return RedirectResponse(f"/sites/{site_id}?mode=insert&err=1", status_code=303)
    ids = [db.add_keyword(site_id, r) for r in rows]
    if run_now:
        jobs.start(site_id, kw_ids=ids)
        return RedirectResponse(f"/sites/{site_id}?mode=insert&running=1", status_code=303)
    return RedirectResponse(f"/sites/{site_id}?mode=insert&added={len(ids)}", status_code=303)


# ---------------- 导出排定任务 ----------------
@app.get("/sites/{site_id}/export")
def export_excel(site_id: int):
    site = db.get_site(site_id)
    rows = db.list_keywords(site_id)
    cols = ["main_keyword", "secondary_keyword", "topic", "wordcounts", "specific",
            "categories", "tags", "status", "word_count", "cost_usd", "wp_link"]
    buf = _rows_to_xlsx(cols, rows)
    fname = f"tasks_site{site_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------- 插入任务模式 ----------------
@app.post("/sites/{site_id}/insert")
async def insert_task(request: Request, site_id: int):
    form = await request.form()
    main_kw = (form.get("main_keyword") or "").strip()
    if not main_kw:
        return RedirectResponse(f"/sites/{site_id}?mode=insert&err=1", status_code=303)
    try:
        wc = int(float(form.get("wordcounts") or 0))
    except (TypeError, ValueError):
        wc = 0
    row = {
        "main_keyword": main_kw,
        "secondary_keyword": (form.get("secondary_keyword") or main_kw).strip(),
        "topic": (form.get("topic") or main_kw).strip(),
        "wordcounts": wc,
        "specific": (form.get("specific") or "").strip(),
        "categories": (form.get("categories") or "").strip(),
        "tags": (form.get("tags") or "").strip(),
    }
    kw_id = db.add_keyword(site_id, row)
    if form.get("run_now"):
        jobs.start(site_id, kw_ids=[kw_id])
        return RedirectResponse(f"/sites/{site_id}?mode=insert&running=1", status_code=303)
    return RedirectResponse(f"/sites/{site_id}?mode=insert&inserted=1", status_code=303)


# ---------------- 关键词行操作 ----------------
@app.post("/sites/{site_id}/keywords/{kw_id}/delete")
def del_keyword(site_id: int, kw_id: int):
    db.delete_keyword(kw_id)
    return RedirectResponse(f"/sites/{site_id}", status_code=303)


@app.post("/sites/{site_id}/keywords/{kw_id}/reset")
def reset_kw(site_id: int, kw_id: int):
    db.reset_keyword(kw_id)
    return RedirectResponse(f"/sites/{site_id}", status_code=303)


# ---------------- 运行控制 ----------------
@app.post("/sites/{site_id}/run")
def run_site(site_id: int, limit: Optional[int] = Form(None)):
    started = jobs.start(site_id, limit=limit)
    return JSONResponse({"ok": started, "msg": "已开始" if started else "已有任务在跑"})


@app.post("/sites/{site_id}/stop")
def stop_site(site_id: int):
    jobs.stop(site_id)
    return JSONResponse({"ok": True})


@app.get("/sites/{site_id}/progress")
def progress(site_id: int):
    job = jobs.get(site_id)
    counts = db.keyword_counts(site_id)
    if not job:
        return JSONResponse({"job": None, "counts": counts})
    return JSONResponse({"job": job.snapshot(), "counts": counts})


# ---------------- 工具 ----------------
def _coerce(values: dict, form) -> dict:
    """把复选框 / 数值字段转成正确类型。"""
    # image_enabled 是复选框：出现即 1
    values["image_enabled"] = 1 if form.get("image_enabled") in ("1", "on", "true") else 0
    for k in db._INT_FIELDS:
        if k == "image_enabled":
            continue
        try:
            values[k] = int(float(values.get(k) or db.SITE_FIELDS[k]))
        except (TypeError, ValueError):
            values[k] = db.SITE_FIELDS[k]
    return values
